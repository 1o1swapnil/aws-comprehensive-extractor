#!/usr/bin/env python3
"""
Comprehensive AWS Information Extraction Script
Extracts detailed information from multiple AWS services
Supports manual AWS profile specification with Excel output
"""

import boto3
import json
from datetime import datetime
from botocore.exceptions import ClientError, NoCredentialsError, ProfileNotFound
import argparse
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import warnings
warnings.filterwarnings('ignore')

class ComprehensiveAWSExtractor:
    def __init__(self, profile_name=None, region='us-east-1'):
        """Initialize AWS session with specified profile"""
        try:
            if profile_name:
                self.session = boto3.Session(profile_name=profile_name)
                print(f"✓ Using AWS profile: {profile_name}")
            else:
                self.session = boto3.Session()
                print("✓ Using default AWS credentials")

            self.region = region
            self.account_id = self._get_account_id()
            self.caller_identity = self._get_caller_identity()
            print(f"✓ Connected to AWS Account: {self.account_id}")
            print(f"✓ Region: {self.region}")

        except ProfileNotFound:
            print(f"❌ Error: AWS profile '{profile_name}' not found")
            self._list_available_profiles()
            sys.exit(1)
        except NoCredentialsError:
            print("❌ Error: No AWS credentials found")
            sys.exit(1)
        except Exception as e:
            print(f"❌ Error initializing AWS session: {e}")
            sys.exit(1)

    def _get_account_id(self):
        """Get current AWS account ID"""
        try:
            sts = self.session.client('sts')
            return sts.get_caller_identity()['Account']
        except Exception as e:
            return "Unknown"

    def _get_caller_identity(self):
        """Get caller identity details"""
        try:
            sts = self.session.client('sts')
            return sts.get_caller_identity()
        except Exception as e:
            return {}

    def _list_available_profiles(self):
        """List available AWS profiles"""
        try:
            session = boto3.Session()
            profiles = session.available_profiles
            print("Available profiles:")
            for profile in profiles:
                print(f"  - {profile}")
        except Exception as e:
            print(f"Error listing profiles: {e}")

    def _safe_extract(self, func, service_name):
        """Safely execute extraction function"""
        try:
            print(f"🔍 Extracting {service_name}...")
            return func()
        except ClientError as e:
            error_code = e.response['Error']['Code']
            if error_code in ['AccessDenied', 'UnauthorizedOperation']:
                print(f"⚠️  {service_name}: Access denied - insufficient permissions")
            else:
                print(f"⚠️  {service_name}: {e}")
            return []
        except Exception as e:
            print(f"⚠️  {service_name}: {e}")
            return []

    def extract_ec2_instances(self):
        """Extract EC2 instances information"""
        ec2 = self.session.client('ec2', region_name=self.region)
        instances = []

        response = ec2.describe_instances()
        for reservation in response['Reservations']:
            for instance in reservation['Instances']:
                instance_info = {
                    'InstanceId': instance.get('InstanceId'),
                    'InstanceType': instance.get('InstanceType'),
                    'State': instance.get('State', {}).get('Name'),
                    'PublicIpAddress': instance.get('PublicIpAddress', 'N/A'),
                    'PrivateIpAddress': instance.get('PrivateIpAddress', 'N/A'),
                    'VpcId': instance.get('VpcId'),
                    'SubnetId': instance.get('SubnetId'),
                    'AvailabilityZone': instance.get('Placement', {}).get('AvailabilityZone'),
                    'SecurityGroups': ', '.join([sg['GroupId'] for sg in instance.get('SecurityGroups', [])]),
                    'KeyName': instance.get('KeyName', 'N/A'),
                    'LaunchTime': str(instance.get('LaunchTime', 'N/A')),
                    'Platform': instance.get('Platform', 'Linux'),
                    'Architecture': instance.get('Architecture'),
                    'Monitoring': instance.get('Monitoring', {}).get('State'),
                    'Tags': self._format_tags(instance.get('Tags', []))
                }
                instances.append(instance_info)

        return instances

    def extract_quicksight(self):
        """Extract QuickSight information"""
        quicksight = self.session.client('quicksight', region_name=self.region)
        data = []

        try:
            # Get data sources
            response = quicksight.list_data_sources(AwsAccountId=self.account_id)
            for ds in response.get('DataSources', []):
                data.append({
                    'Type': 'DataSource',
                    'Name': ds.get('Name'),
                    'DataSourceId': ds.get('DataSourceId'),
                    'Type_Detail': ds.get('Type'),
                    'Status': ds.get('Status'),
                    'CreatedTime': str(ds.get('CreatedTime', 'N/A')),
                    'LastUpdatedTime': str(ds.get('LastUpdatedTime', 'N/A'))
                })
        except Exception as e:
            print(f"QuickSight DataSources error: {e}")

        try:
            # Get dashboards
            response = quicksight.list_dashboards(AwsAccountId=self.account_id)
            for dashboard in response.get('DashboardSummaryList', []):
                data.append({
                    'Type': 'Dashboard',
                    'Name': dashboard.get('Name'),
                    'DashboardId': dashboard.get('DashboardId'),
                    'PublishedVersionNumber': dashboard.get('PublishedVersionNumber'),
                    'CreatedTime': str(dashboard.get('CreatedTime', 'N/A')),
                    'LastUpdatedTime': str(dashboard.get('LastUpdatedTime', 'N/A')),
                    'LastPublishedTime': str(dashboard.get('LastPublishedTime', 'N/A'))
                })
        except Exception as e:
            print(f"QuickSight Dashboards error: {e}")

        return data

    def extract_load_balancers(self):
        """Extract Elastic Load Balancer information"""
        # Application Load Balancers
        elbv2 = self.session.client('elbv2', region_name=self.region)
        albs = []

        try:
            response = elbv2.describe_load_balancers()
            for lb in response['LoadBalancers']:
                alb_info = {
                    'Type': 'ALB/NLB',
                    'LoadBalancerName': lb.get('LoadBalancerName'),
                    'DNSName': lb.get('DNSName'),
                    'LoadBalancerType': lb.get('Type'),
                    'Scheme': lb.get('Scheme'),
                    'State': lb.get('State', {}).get('Code'),
                    'VpcId': lb.get('VpcId'),
                    'AvailabilityZones': ', '.join([az['ZoneName'] for az in lb.get('AvailabilityZones', [])]),
                    'SecurityGroups': ', '.join(lb.get('SecurityGroups', [])),
                    'CreatedTime': str(lb.get('CreatedTime', 'N/A'))
                }
                albs.append(alb_info)
        except Exception as e:
            print(f"ALB/NLB extraction error: {e}")

        # Classic Load Balancers
        elb = self.session.client('elb', region_name=self.region)
        try:
            response = elb.describe_load_balancers()
            for lb in response['LoadBalancerDescriptions']:
                clb_info = {
                    'Type': 'CLB',
                    'LoadBalancerName': lb.get('LoadBalancerName'),
                    'DNSName': lb.get('DNSName'),
                    'LoadBalancerType': 'classic',
                    'Scheme': lb.get('Scheme'),
                    'VpcId': lb.get('VPCId'),
                    'AvailabilityZones': ', '.join(lb.get('AvailabilityZones', [])),
                    'SecurityGroups': ', '.join(lb.get('SecurityGroups', [])),
                    'CreatedTime': str(lb.get('CreatedTime', 'N/A'))
                }
                albs.append(clb_info)
        except Exception as e:
            print(f"CLB extraction error: {e}")

        return albs

    def extract_ecs(self):
        """Extract ECS information"""
        ecs = self.session.client('ecs', region_name=self.region)
        data = []

        # Get clusters
        try:
            clusters_response = ecs.list_clusters()
            for cluster_arn in clusters_response['clusterArns']:
                cluster_details = ecs.describe_clusters(clusters=[cluster_arn])
                for cluster in cluster_details['clusters']:
                    data.append({
                        'Type': 'Cluster',
                        'Name': cluster.get('clusterName'),
                        'Status': cluster.get('status'),
                        'RunningTasksCount': cluster.get('runningTasksCount'),
                        'PendingTasksCount': cluster.get('pendingTasksCount'),
                        'ActiveServicesCount': cluster.get('activeServicesCount'),
                        'RegisteredContainerInstancesCount': cluster.get('registeredContainerInstancesCount'),
                        'Tags': self._format_tags(cluster.get('tags', []))
                    })
        except Exception as e:
            print(f"ECS Clusters error: {e}")

        return data

    def extract_documentdb(self):
        """Extract DocumentDB information"""
        docdb = self.session.client('docdb', region_name=self.region)
        data = []

        try:
            # Get clusters
            response = docdb.describe_db_clusters()
            for cluster in response['DBClusters']:
                data.append({
                    'Type': 'Cluster',
                    'DBClusterIdentifier': cluster.get('DBClusterIdentifier'),
                    'Engine': cluster.get('Engine'),
                    'EngineVersion': cluster.get('EngineVersion'),
                    'Status': cluster.get('Status'),
                    'Endpoint': cluster.get('Endpoint'),
                    'Port': cluster.get('Port'),
                    'MasterUsername': cluster.get('MasterUsername'),
                    'BackupRetentionPeriod': cluster.get('BackupRetentionPeriod'),
                    'VpcSecurityGroups': ', '.join([sg['VpcSecurityGroupId'] for sg in cluster.get('VpcSecurityGroups', [])]),
                    'DBSubnetGroup': cluster.get('DBSubnetGroup'),
                    'ClusterCreateTime': str(cluster.get('ClusterCreateTime', 'N/A'))
                })
        except Exception as e:
            print(f"DocumentDB error: {e}")

        return data

    def extract_eks(self):
        """Extract EKS information"""
        eks = self.session.client('eks', region_name=self.region)
        data = []

        try:
            # Get clusters
            response = eks.list_clusters()
            for cluster_name in response['clusters']:
                cluster_details = eks.describe_cluster(name=cluster_name)
                cluster = cluster_details['cluster']

                data.append({
                    'Type': 'Cluster',
                    'Name': cluster.get('name'),
                    'Status': cluster.get('status'),
                    'Version': cluster.get('version'),
                    'Endpoint': cluster.get('endpoint'),
                    'RoleArn': cluster.get('roleArn'),
                    'VpcConfig': str(cluster.get('resourcesVpcConfig', {})),
                    'CreatedAt': str(cluster.get('createdAt', 'N/A')),
                    'Tags': self._format_tags(cluster.get('tags', {}))
                })
        except Exception as e:
            print(f"EKS error: {e}")

        return data

    def extract_textract(self):
        """Extract Textract information (jobs and usage)"""
        # Textract doesn't have list operations, so we'll return service availability info
        textract = self.session.client('textract', region_name=self.region)
        data = []

        try:
            # Check if service is available by making a simple call
            textract.get_document_analysis(JobId='test')  # This will fail but tells us service is available
        except ClientError as e:
            if 'InvalidJobId' in str(e):
                data.append({
                    'Service': 'Textract',
                    'Status': 'Available',
                    'Region': self.region,
                    'Note': 'Service is available - no active jobs to list'
                })
            else:
                data.append({
                    'Service': 'Textract',
                    'Status': 'Error',
                    'Region': self.region,
                    'Error': str(e)
                })
        except Exception as e:
            data.append({
                'Service': 'Textract',
                'Status': 'Unknown',
                'Region': self.region,
                'Error': str(e)
            })

        return data

    def extract_rekognition(self):
        """Extract Rekognition information"""
        rekognition = self.session.client('rekognition', region_name=self.region)
        data = []

        try:
            # Get collections
            response = rekognition.list_collections()
            for collection_id in response['CollectionIds']:
                collection_info = rekognition.describe_collection(CollectionId=collection_id)
                data.append({
                    'Type': 'Collection',
                    'CollectionId': collection_id,
                    'FaceCount': collection_info.get('FaceCount'),
                    'FaceModelVersion': collection_info.get('FaceModelVersion'),
                    'CollectionARN': collection_info.get('CollectionARN'),
                    'CreationTimestamp': str(collection_info.get('CreationTimestamp', 'N/A'))
                })
        except Exception as e:
            print(f"Rekognition error: {e}")
            data.append({
                'Service': 'Rekognition',
                'Status': 'Available',
                'Region': self.region,
                'Note': 'Service is available - no collections found'
            })

        return data

    def extract_kms(self):
        """Extract KMS information"""
        kms = self.session.client('kms', region_name=self.region)
        data = []

        try:
            # Get keys
            response = kms.list_keys()
            for key in response['Keys']:
                key_details = kms.describe_key(KeyId=key['KeyId'])
                key_metadata = key_details['KeyMetadata']

                data.append({
                    'KeyId': key_metadata.get('KeyId'),
                    'KeyArn': key_metadata.get('Arn'),
                    'Description': key_metadata.get('Description', 'N/A'),
                    'KeyUsage': key_metadata.get('KeyUsage'),
                    'KeyState': key_metadata.get('KeyState'),
                    'Origin': key_metadata.get('Origin'),
                    'KeyManager': key_metadata.get('KeyManager'),
                    'CreationDate': str(key_metadata.get('CreationDate', 'N/A')),
                    'Enabled': key_metadata.get('Enabled')
                })
        except Exception as e:
            print(f"KMS error: {e}")

        return data

    def extract_bedrock(self):
        """Extract Bedrock information"""
        try:
            bedrock = self.session.client('bedrock', region_name=self.region)
            data = []

            # Get foundation models
            response = bedrock.list_foundation_models()
            for model in response.get('modelSummaries', []):
                data.append({
                    'Type': 'Foundation Model',
                    'ModelId': model.get('modelId'),
                    'ModelName': model.get('modelName'),
                    'ProviderName': model.get('providerName'),
                    'InputModalities': ', '.join(model.get('inputModalities', [])),
                    'OutputModalities': ', '.join(model.get('outputModalities', [])),
                    'ResponseStreamingSupported': model.get('responseStreamingSupported'),
                    'CustomizationsSupported': ', '.join(model.get('customizationsSupported', []))
                })
        except Exception as e:
            print(f"Bedrock error: {e}")
            data = [{'Service': 'Bedrock', 'Status': 'Not available in region or access denied', 'Error': str(e)}]

        return data

    def extract_cloudtrail(self):
        """Extract CloudTrail information"""
        cloudtrail = self.session.client('cloudtrail', region_name=self.region)
        data = []

        try:
            # Get trails
            response = cloudtrail.describe_trails()
            for trail in response['trailList']:
                trail_status = cloudtrail.get_trail_status(Name=trail['TrailARN'])

                data.append({
                    'Name': trail.get('Name'),
                    'S3BucketName': trail.get('S3BucketName'),
                    'IncludeGlobalServiceEvents': trail.get('IncludeGlobalServiceEvents'),
                    'IsMultiRegionTrail': trail.get('IsMultiRegionTrail'),
                    'IsLogging': trail_status.get('IsLogging'),
                    'LogFileValidationEnabled': trail.get('LogFileValidationEnabled'),
                    'EventSelectors': 'Configured' if trail.get('HasCustomEventSelectors') else 'Default',
                    'KMSKeyId': trail.get('KMSKeyId', 'N/A'),
                    'TrailARN': trail.get('TrailARN')
                })
        except Exception as e:
            print(f"CloudTrail error: {e}")

        return data

    def extract_systems_manager(self):
        """Extract Systems Manager information"""
        ssm = self.session.client('ssm', region_name=self.region)
        data = []

        try:
            # Get managed instances
            response = ssm.describe_instance_information()
            for instance in response['InstanceInformationList']:
                data.append({
                    'Type': 'Managed Instance',
                    'InstanceId': instance.get('InstanceId'),
                    'PlatformType': instance.get('PlatformType'),
                    'PlatformName': instance.get('PlatformName'),
                    'PlatformVersion': instance.get('PlatformVersion'),
                    'AgentVersion': instance.get('AgentVersion'),
                    'PingStatus': instance.get('PingStatus'),
                    'LastPingDateTime': str(instance.get('LastPingDateTime', 'N/A')),
                    'IPAddress': instance.get('IPAddress')
                })
        except Exception as e:
            print(f"Systems Manager error: {e}")

        return data

    def extract_guardduty(self):
        """Extract GuardDuty information"""
        guardduty = self.session.client('guardduty', region_name=self.region)
        data = []

        try:
            # Get detectors
            response = guardduty.list_detectors()
            for detector_id in response['DetectorIds']:
                detector_details = guardduty.get_detector(DetectorId=detector_id)

                data.append({
                    'DetectorId': detector_id,
                    'Status': detector_details.get('Status'),
                    'ServiceRole': detector_details.get('ServiceRole'),
                    'FindingPublishingFrequency': detector_details.get('FindingPublishingFrequency'),
                    'CreatedAt': str(detector_details.get('CreatedAt', 'N/A')),
                    'UpdatedAt': str(detector_details.get('UpdatedAt', 'N/A')),
                    'Tags': self._format_tags(detector_details.get('Tags', {}))
                })
        except Exception as e:
            print(f"GuardDuty error: {e}")

        return data

    def extract_api_gateway(self):
        """Extract API Gateway information"""
        # REST APIs
        apigateway = self.session.client('apigateway', region_name=self.region)
        data = []

        try:
            response = apigateway.get_rest_apis()
            for api in response['items']:
                data.append({
                    'Type': 'REST API',
                    'Name': api.get('name'),
                    'Id': api.get('id'),
                    'Description': api.get('description', 'N/A'),
                    'CreatedDate': str(api.get('createdDate', 'N/A')),
                    'EndpointConfiguration': str(api.get('endpointConfiguration', {})),
                    'Tags': self._format_tags(api.get('tags', {}))
                })
        except Exception as e:
            print(f"API Gateway REST error: {e}")

        # HTTP APIs (API Gateway v2)
        try:
            apigatewayv2 = self.session.client('apigatewayv2', region_name=self.region)
            response = apigatewayv2.get_apis()
            for api in response['Items']:
                data.append({
                    'Type': 'HTTP API',
                    'Name': api.get('Name'),
                    'ApiId': api.get('ApiId'),
                    'Description': api.get('Description', 'N/A'),
                    'CreatedDate': str(api.get('CreatedDate', 'N/A')),
                    'ProtocolType': api.get('ProtocolType'),
                    'Tags': self._format_tags(api.get('Tags', {}))
                })
        except Exception as e:
            print(f"API Gateway HTTP error: {e}")

        return data

    def extract_cloudformation(self):
        """Extract CloudFormation information"""
        cf = self.session.client('cloudformation', region_name=self.region)
        data = []

        try:
            response = cf.describe_stacks()
            for stack in response['Stacks']:
                data.append({
                    'StackName': stack.get('StackName'),
                    'StackStatus': stack.get('StackStatus'),
                    'CreationTime': str(stack.get('CreationTime', 'N/A')),
                    'LastUpdatedTime': str(stack.get('LastUpdatedTime', 'N/A')),
                    'Description': stack.get('Description', 'N/A'),
                    'Parameters': len(stack.get('Parameters', [])),
                    'Outputs': len(stack.get('Outputs', [])),
                    'Tags': self._format_tags(stack.get('Tags', []))
                })
        except Exception as e:
            print(f"CloudFormation error: {e}")

        return data

    def extract_location_service(self):
        """Extract Location Service information"""
        location = self.session.client('location', region_name=self.region)
        data = []

        try:
            # Get maps
            response = location.list_maps()
            for map_resource in response['Entries']:
                data.append({
                    'Type': 'Map',
                    'MapName': map_resource.get('MapName'),
                    'Description': map_resource.get('Description', 'N/A'),
                    'DataSource': map_resource.get('DataSource'),
                    'CreateTime': str(map_resource.get('CreateTime', 'N/A')),
                    'UpdateTime': str(map_resource.get('UpdateTime', 'N/A'))
                })
        except Exception as e:
            print(f"Location Service error: {e}")

        return data

    def extract_neptune(self):
        """Extract Neptune information"""
        neptune = self.session.client('neptune', region_name=self.region)
        data = []

        try:
            # Get clusters
            response = neptune.describe_db_clusters()
            for cluster in response['DBClusters']:
                data.append({
                    'Type': 'Cluster',
                    'DBClusterIdentifier': cluster.get('DBClusterIdentifier'),
                    'Engine': cluster.get('Engine'),
                    'EngineVersion': cluster.get('EngineVersion'),
                    'Status': cluster.get('Status'),
                    'Endpoint': cluster.get('Endpoint'),
                    'Port': cluster.get('Port'),
                    'MasterUsername': cluster.get('MasterUsername'),
                    'BackupRetentionPeriod': cluster.get('BackupRetentionPeriod'),
                    'ClusterCreateTime': str(cluster.get('ClusterCreateTime', 'N/A'))
                })
        except Exception as e:
            print(f"Neptune error: {e}")

        return data

    def extract_service_catalog(self):
        """Extract Service Catalog information"""
        sc = self.session.client('servicecatalog', region_name=self.region)
        data = []

        try:
            # Get portfolios
            response = sc.list_portfolios()
            for portfolio in response['PortfolioDetails']:
                data.append({
                    'Type': 'Portfolio',
                    'Id': portfolio.get('Id'),
                    'ARN': portfolio.get('ARN'),
                    'DisplayName': portfolio.get('DisplayName'),
                    'Description': portfolio.get('Description', 'N/A'),
                    'ProviderName': portfolio.get('ProviderName'),
                    'CreatedTime': str(portfolio.get('CreatedTime', 'N/A'))
                })
        except Exception as e:
            print(f"Service Catalog error: {e}")

        return data

    def extract_glue(self):
        """Extract AWS Glue information"""
        glue = self.session.client('glue', region_name=self.region)
        data = []

        try:
            # Get databases
            response = glue.get_databases()
            for database in response['DatabaseList']:
                data.append({
                    'Type': 'Database',
                    'Name': database.get('Name'),
                    'Description': database.get('Description', 'N/A'),
                    'CreateTime': str(database.get('CreateTime', 'N/A')),
                    'Parameters': str(database.get('Parameters', {}))
                })
        except Exception as e:
            print(f"Glue error: {e}")

        return data

    def extract_codeartifact(self):
        """Extract CodeArtifact information"""
        codeartifact = self.session.client('codeartifact', region_name=self.region)
        data = []

        try:
            # Get domains
            response = codeartifact.list_domains()
            for domain in response['domains']:
                data.append({
                    'Type': 'Domain',
                    'Name': domain.get('name'),
                    'Owner': domain.get('owner'),
                    'Arn': domain.get('arn'),
                    'Status': domain.get('status'),
                    'CreatedTime': str(domain.get('createdTime', 'N/A'))
                })
        except Exception as e:
            print(f"CodeArtifact error: {e}")

        return data

    def extract_dynamodb(self):
        """Extract DynamoDB information"""
        dynamodb = self.session.client('dynamodb', region_name=self.region)
        data = []

        try:
            # Get tables
            response = dynamodb.list_tables()
            for table_name in response['TableNames']:
                table_details = dynamodb.describe_table(TableName=table_name)
                table = table_details['Table']

                data.append({
                    'TableName': table.get('TableName'),
                    'TableStatus': table.get('TableStatus'),
                    'CreationDateTime': str(table.get('CreationDateTime', 'N/A')),
                    'ItemCount': table.get('ItemCount'),
                    'TableSizeBytes': table.get('TableSizeBytes'),
                    'BillingMode': table.get('BillingModeSummary', {}).get('BillingMode'),
                    'GlobalSecondaryIndexes': len(table.get('GlobalSecondaryIndexes', [])),
                    'LocalSecondaryIndexes': len(table.get('LocalSecondaryIndexes', [])),
                    'StreamSpecification': str(table.get('StreamSpecification', {}))
                })
        except Exception as e:
            print(f"DynamoDB error: {e}")

        return data

    def extract_memorydb(self):
        """Extract MemoryDB information"""
        memorydb = self.session.client('memorydb', region_name=self.region)
        data = []

        try:
            # Get clusters
            response = memorydb.describe_clusters()
            for cluster in response['Clusters']:
                data.append({
                    'Name': cluster.get('Name'),
                    'Status': cluster.get('Status'),
                    'NodeType': cluster.get('NodeType'),
                    'Engine': cluster.get('Engine'),
                    'EngineVersion': cluster.get('EngineVersion'),
                    'NumShards': cluster.get('NumShards'),
                    'SecurityGroups': ', '.join([sg['SecurityGroupId'] for sg in cluster.get('SecurityGroups', [])]),
                    'SubnetGroupName': cluster.get('SubnetGroupName'),
                    'TLSEnabled': cluster.get('TLSEnabled')
                })
        except Exception as e:
            print(f"MemoryDB error: {e}")

        return data

    def extract_vpc(self):
        """Extract VPC information"""
        ec2 = self.session.client('ec2', region_name=self.region)
        data = []

        try:
            # Get VPCs
            response = ec2.describe_vpcs()
            for vpc in response['Vpcs']:
                data.append({
                    'Type': 'VPC',
                    'VpcId': vpc.get('VpcId'),
                    'CidrBlock': vpc.get('CidrBlock'),
                    'State': vpc.get('State'),
                    'IsDefault': vpc.get('IsDefault'),
                    'DhcpOptionsId': vpc.get('DhcpOptionsId'),
                    'InstanceTenancy': vpc.get('InstanceTenancy'),
                    'Tags': self._format_tags(vpc.get('Tags', []))
                })

            # Get Subnets
            response = ec2.describe_subnets()
            for subnet in response['Subnets']:
                data.append({
                    'Type': 'Subnet',
                    'SubnetId': subnet.get('SubnetId'),
                    'VpcId': subnet.get('VpcId'),
                    'CidrBlock': subnet.get('CidrBlock'),
                    'AvailabilityZone': subnet.get('AvailabilityZone'),
                    'State': subnet.get('State'),
                    'MapPublicIpOnLaunch': subnet.get('MapPublicIpOnLaunch'),
                    'Tags': self._format_tags(subnet.get('Tags', []))
                })
        except Exception as e:
            print(f"VPC error: {e}")

        return data

    def extract_sagemaker(self):
        """Extract SageMaker information"""
        sagemaker = self.session.client('sagemaker', region_name=self.region)
        data = []

        try:
            # Get notebook instances
            response = sagemaker.list_notebook_instances()
            for notebook in response['NotebookInstances']:
                data.append({
                    'Type': 'Notebook Instance',
                    'NotebookInstanceName': notebook.get('NotebookInstanceName'),
                    'NotebookInstanceStatus': notebook.get('NotebookInstanceStatus'),
                    'InstanceType': notebook.get('InstanceType'),
                    'CreationTime': str(notebook.get('CreationTime', 'N/A')),
                    'LastModifiedTime': str(notebook.get('LastModifiedTime', 'N/A')),
                    'Url': notebook.get('Url', 'N/A')
                })
        except Exception as e:
            print(f"SageMaker error: {e}")

        return data

    def extract_efs(self):
        """Extract EFS information"""
        efs = self.session.client('efs', region_name=self.region)
        data = []

        try:
            # Get file systems
            response = efs.describe_file_systems()
            for fs in response['FileSystems']:
                data.append({
                    'FileSystemId': fs.get('FileSystemId'),
                    'Name': fs.get('Name', 'N/A'),
                    'CreationTime': str(fs.get('CreationTime', 'N/A')),
                    'LifeCycleState': fs.get('LifeCycleState'),
                    'NumberOfMountTargets': fs.get('NumberOfMountTargets'),
                    'SizeInBytes': fs.get('SizeInBytes', {}).get('Value'),
                    'PerformanceMode': fs.get('PerformanceMode'),
                    'ThroughputMode': fs.get('ThroughputMode'),
                    'Encrypted': fs.get('Encrypted'),
                    'Tags': self._format_tags(fs.get('Tags', []))
                })
        except Exception as e:
            print(f"EFS error: {e}")

        return data

    def extract_opensearch(self):
        """Extract OpenSearch information"""
        opensearch = self.session.client('opensearch', region_name=self.region)
        data = []

        try:
            # Get domains
            response = opensearch.list_domain_names()
            for domain in response['DomainNames']:
                domain_name = domain['DomainName']
                domain_details = opensearch.describe_domain(DomainName=domain_name)
                domain_status = domain_details['DomainStatus']

                data.append({
                    'DomainName': domain_status.get('DomainName'),
                    'ElasticsearchVersion': domain_status.get('ElasticsearchVersion'),
                    'Created': domain_status.get('Created'),
                    'Deleted': domain_status.get('Deleted'),
                    'Endpoint': domain_status.get('Endpoint'),
                    'Processing': domain_status.get('Processing'),
                    'UpgradeProcessing': domain_status.get('UpgradeProcessing'),
                    'InstanceType': domain_status.get('ClusterConfig', {}).get('InstanceType'),
                    'InstanceCount': domain_status.get('ClusterConfig', {}).get('InstanceCount')
                })
        except Exception as e:
            print(f"OpenSearch error: {e}")

        return data

    def extract_security_hub(self):
        """Extract Security Hub information"""
        securityhub = self.session.client('securityhub', region_name=self.region)
        data = []

        try:
            # Get hub details
            response = securityhub.describe_hub()
            data.append({
                'HubArn': response.get('HubArn'),
                'SubscribedAt': str(response.get('SubscribedAt', 'N/A')),
                'AutoEnableControls': response.get('AutoEnableControls')
            })
        except Exception as e:
            print(f"Security Hub error: {e}")
            data.append({
                'Service': 'Security Hub',
                'Status': 'Not enabled or access denied',
                'Error': str(e)
            })

        return data

    def extract_xray(self):
        """Extract X-Ray information"""
        xray = self.session.client('xray', region_name=self.region)
        data = []

        try:
            # Get sampling rules
            response = xray.get_sampling_rules()
            for rule in response['SamplingRuleRecords']:
                sampling_rule = rule['SamplingRule']
                data.append({
                    'Type': 'Sampling Rule',
                    'RuleName': sampling_rule.get('RuleName'),
                    'Priority': sampling_rule.get('Priority'),
                    'FixedRate': sampling_rule.get('FixedRate'),
                    'ReservoirSize': sampling_rule.get('ReservoirSize'),
                    'ServiceName': sampling_rule.get('ServiceName'),
                    'ServiceType': sampling_rule.get('ServiceType'),
                    'Host': sampling_rule.get('Host'),
                    'HTTPMethod': sampling_rule.get('HTTPMethod'),
                    'URLPath': sampling_rule.get('URLPath'),
                    'Version': sampling_rule.get('Version')
                })
        except Exception as e:
            print(f"X-Ray error: {e}")

        return data

    def _format_tags(self, tags):
        """Format tags for display"""
        if not tags:
            return "No tags"
        if isinstance(tags, dict):
            return "; ".join([f"{k}={v}" for k, v in tags.items()])
        if isinstance(tags, list):
            return "; ".join([f"{tag.get('Key', '')}={tag.get('Value', '')}" for tag in tags])
        return str(tags)

    def create_excel_report(self, all_data):
        """Create comprehensive Excel report"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"AWS_Comprehensive_Report_{timestamp}.xlsx"

        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Create summary sheet
        summary_ws = wb.create_sheet("Summary")
        summary_data = [
            ["AWS Comprehensive Report", ""],
            ["Account ID", self.account_id],
            ["Region", self.region],
            ["User/Role ARN", self.caller_identity.get('Arn', 'Unknown')],
            ["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["", ""],
            ["Service", "Total Resources"]
        ]

        # Add service counts to summary
        for service_name, data in all_data.items():
            if isinstance(data, list):
                summary_data.append([service_name, len(data)])

        for row_num, row_data in enumerate(summary_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = summary_ws.cell(row=row_num, column=col_num, value=value)
                if row_num == 1 or row_num == 7:  # Header rows
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

        # Create sheets for each service
        for service_name, data in all_data.items():
            if not data:
                continue

            ws = wb.create_sheet(service_name)

            if isinstance(data, list) and len(data) > 0:
                # Convert to DataFrame for better handling
                df = pd.DataFrame(data)

                # Write headers
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = ws.cell(row=1, column=col_num, value=column_title)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                # Write data
                for row_num, row_data in enumerate(df.values, 2):
                    for col_num, value in enumerate(row_data, 1):
                        # Handle complex data types
                        if isinstance(value, (list, dict)):
                            value = str(value)
                        ws.cell(row=row_num, column=col_num, value=value)

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    ws.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        wb.save(filename)
        print(f"\n✅ Comprehensive Excel report saved: {filename}")
        return filename

    def extract_all_services(self):
        """Extract information from all AWS services"""
        print("\n🚀 Starting comprehensive AWS extraction...")
        print("=" * 60)

        all_data = {}

        # Define all extraction methods
        extractions = [
            ("EC2_Instances", self.extract_ec2_instances),
            ("QuickSight", self.extract_quicksight),
            ("Load_Balancers", self.extract_load_balancers),
            ("ECS", self.extract_ecs),
            ("DocumentDB", self.extract_documentdb),
            ("EKS", self.extract_eks),
            ("Textract", self.extract_textract),
            ("Rekognition", self.extract_rekognition),
            ("KMS", self.extract_kms),
            ("Bedrock", self.extract_bedrock),
            ("CloudTrail", self.extract_cloudtrail),
            ("Systems_Manager", self.extract_systems_manager),
            ("GuardDuty", self.extract_guardduty),
            ("API_Gateway", self.extract_api_gateway),
            ("CloudFormation", self.extract_cloudformation),
            ("Location_Service", self.extract_location_service),
            ("Neptune", self.extract_neptune),
            ("Service_Catalog", self.extract_service_catalog),
            ("Glue", self.extract_glue),
            ("CodeArtifact", self.extract_codeartifact),
            ("DynamoDB", self.extract_dynamodb),
            ("MemoryDB", self.extract_memorydb),
            ("VPC", self.extract_vpc),
            ("SageMaker", self.extract_sagemaker),
            ("EFS", self.extract_efs),
            ("OpenSearch", self.extract_opensearch),
            ("Security_Hub", self.extract_security_hub),
            ("X_Ray", self.extract_xray)
        ]

        # Execute all extractions
        for service_name, extraction_func in extractions:
            all_data[service_name] = self._safe_extract(extraction_func, service_name)

        # Create Excel report
        self.create_excel_report(all_data)

        # Print summary
        print("\n📊 Extraction Summary:")
        print("-" * 40)
        total_resources = 0
        for service_name, data in all_data.items():
            count = len(data) if isinstance(data, list) else 0
            total_resources += count
            print(f"{service_name:<20}: {count:>5} resources")

        print("-" * 40)
        print(f"{'Total Resources':<20}: {total_resources:>5}")
        print("\n✅ Comprehensive extraction completed!")

def main():
    parser = argparse.ArgumentParser(description='Comprehensive AWS Information Extraction Tool')
    parser.add_argument('--profile', '-p', help='AWS profile name')
    parser.add_argument('--region', '-r', default='us-east-1', help='AWS region (default: us-east-1)')

    args = parser.parse_args()

    print("🚀 AWS Comprehensive Information Extraction Tool")
    print("=" * 60)
    print("📋 This tool will extract information from all specified AWS services")
    print("📊 Output will be generated in Excel format with multiple sheets")
    print("⚠️  Note: Some services may require specific permissions")

    try:
        extractor = ComprehensiveAWSExtractor(profile_name=args.profile, region=args.region)
        extractor.extract_all_services()

    except KeyboardInterrupt:
        print("\n\n👋 Script interrupted by user")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ Unexpected error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
